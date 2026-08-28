#!/usr/bin/env python3
"""スプレッドシート（Excel/CSV）の会議情報・議題を議事次第メールテンプレートに流し込む。

使い方:
    # Excelワークブック（シート「会議情報」「議題」を含む .xlsx）から生成
    python3 scripts/generate_agenda_email.py --workbook meeting_data.xlsx

    # CSV2ファイル（会議情報・議題を分けて管理）から生成
    python3 scripts/generate_agenda_email.py --info meeting_info.csv --items agenda_items.csv

出力先は省略時 output/agenda_<年>-<月>.html。
ブラウザで開いて全選択→コピーし、メールソフトの本文に貼り付けて送信する。

入力フォーマットは scripts/sample_data/ のサンプルを参照。
"""
import argparse
import csv
import datetime
import html
import re
import sys
from pathlib import Path

SCRIPT_VERSION = "2026-08-28"

REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_TEMPLATE = REPO_ROOT / "meeting_agenda_template.html"

REQUIRED_INFO_KEYS = ["年", "月日", "時間帯", "場所", "議長", "記録係", "出席者"]
OPTIONAL_INFO_KEYS = ["欠席者", "件名"]
# 「月日」の項目名として区切り文字だけ違う表記（月/日、月ノ日、月・日 など）を許容する。
MONTH_DAY_SEPARATORS = "/\\-ノ・　 "

# 議題シートの列名。「説明・論点」「資料/備考」は空欄可。
AGENDA_REQUIRED_COLUMNS = ["No", "議題", "担当", "想定時間（分）"]
AGENDA_OPTIONAL_COLUMNS = ["説明・論点", "資料/備考"]
AGENDA_COLUMNS = AGENDA_REQUIRED_COLUMNS + AGENDA_OPTIONAL_COLUMNS

WEEKDAY_JA = ["月", "火", "水", "木", "金", "土", "日"]

ROW_TEMPLATE = """        <tr>
          <td align="center" style="font-size:13px; color:#1a1a1a; padding:12px 6px;{border}">{no}</td>
          <td style="font-size:13px; color:#1a1a1a; padding:12px 10px;{border}">
            {topic}{sub_lines}
          </td>
          <td style="font-size:13px; color:#1a1a1a; padding:12px 10px;{border}">{owner}</td>
          <td align="center" style="font-size:13px; color:#1a1a1a; padding:12px 6px;{border}">{minutes}分</td>
        </tr>"""
ROW_BORDER = " border-bottom:1px solid #eef0f2;"
SUB_LINE = '<br><span style="font-size:11px; color:#8a92a0; line-height:1.5;">{label}{text}</span>'


def read_info_csv(path):
    info = {}
    with open(path, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if not row.get("項目"):
                continue
            info[row["項目"].strip()] = (row.get("内容") or "").strip()
    return info


def read_items_csv(path):
    items = []
    with open(path, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if not any((row.get(k) or "").strip() for k in AGENDA_COLUMNS):
                continue
            items.append({k: (row.get(k) or "").strip() for k in AGENDA_COLUMNS})
    return items


def cell_to_str(value):
    """Excelセルの値を文字列化する。

    - 整数値扱いの小数（5.0など）は末尾の.0を除く。
    - 「8/28」のように入力してExcelが自動で日付型に変換したセルは
      "8/28" 形式のテキストに戻す（parse_month_dayが読める形にする）。
    """
    if value is None:
        return ""
    if isinstance(value, (datetime.datetime, datetime.date)):
        return f"{value.month}/{value.day}"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def read_workbook(path):
    try:
        import openpyxl
    except ImportError:
        sys.exit(
            "xlsxファイルを読むには openpyxl が必要です。`pip install openpyxl` を実行してください。"
        )
    wb = openpyxl.load_workbook(path, data_only=True)
    if "会議情報" not in wb.sheetnames:
        sys.exit('シート「会議情報」が見つかりません。')
    if "議題" not in wb.sheetnames:
        sys.exit('シート「議題」が見つかりません。')

    info_ws = wb["会議情報"]
    info = {}
    for row in info_ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        key, value = row[0], row[1] if len(row) > 1 else None
        info[cell_to_str(key)] = cell_to_str(value)

    items_ws = wb["議題"]
    header = None
    header_row_index = None
    scan_limit = min(20, items_ws.max_row)
    for row_num, row in enumerate(
        items_ws.iter_rows(min_row=1, max_row=scan_limit, values_only=True), start=1
    ):
        if row and cell_to_str(row[0]) == "No":
            header = [cell_to_str(c) for c in row]
            header_row_index = row_num
            break
    if header is None:
        sys.exit(
            'シート「議題」で見出し行（A列が「No」の行）が見つかりません。'
            '見出し行の1列目のセルが厳密に「No」になっているか確認してください。'
        )
    items = []
    for row in items_ws.iter_rows(min_row=header_row_index + 1, values_only=True):
        if not row or all(v is None for v in row):
            continue
        record = {
            header[i]: cell_to_str(row[i]) for i in range(min(len(header), len(row)))
        }
        if not any(record.get(k) for k in AGENDA_COLUMNS):
            continue
        items.append({k: record.get(k, "") for k in AGENDA_COLUMNS})
    return info, items


def strip_separators(text):
    for ch in MONTH_DAY_SEPARATORS:
        text = text.replace(ch, "")
    return text


def normalize_info_keys(info):
    """「月/日」「月ノ日」のように区切り文字だけ違う項目名を「月日」に寄せる。"""
    if "月日" in info:
        return info
    for key in info:
        if strip_separators(key) == "月日":
            info["月日"] = info[key]
            break
    return info


def validate_info(info):
    missing = [k for k in REQUIRED_INFO_KEYS if not info.get(k)]
    if missing:
        sys.exit(f"会議情報に不足があります: {', '.join(missing)}")


def parse_month_day(value):
    """「9/10」のような「月/日」形式の文字列から (月, 日) を取り出す。"""
    text = str(value).strip()
    for sep in ("/", "-"):
        if sep in text:
            month_str, _, day_str = text.partition(sep)
            try:
                return int(month_str), int(day_str)
            except ValueError:
                break
    sys.exit(f"「月日」の値が正しくありません: {value!r}。「9/10」のように「月/日」の形式で入力してください。")


def format_datetime(info, month, day):
    try:
        year = int(info["年"])
        d = datetime.date(year, month, day)
    except (KeyError, ValueError, TypeError):
        sys.exit(f"「年」の値が正しくありません（年={info.get('年')!r}）。")
    weekday = WEEKDAY_JA[d.weekday()]
    return f"{year}年{month}月{day}日（{weekday}）　{info['時間帯']}"


def replace_field(text, key, value):
    pattern = re.compile(r"(<!--F:%s-->)(.*?)(<!--/F-->)" % re.escape(key), re.DOTALL)
    if not pattern.search(text):
        sys.exit(f"テンプレートにマーカー <!--F:{key}--> が見つかりません。")
    return pattern.sub(lambda m: m.group(1) + html.escape(value) + m.group(3), text)


def replace_rows(text, marker, rows_html):
    pattern = re.compile(
        r"(<!--%s:START-->)(.*?)(<!--%s:END-->)" % (re.escape(marker), re.escape(marker)),
        re.DOTALL,
    )
    if not pattern.search(text):
        sys.exit(f"テンプレートにマーカー <!--{marker}:START--> が見つかりません。")
    return pattern.sub(lambda m: m.group(1) + "\n" + rows_html + "\n" + m.group(3), text)


def build_agenda_rows(items):
    if not items:
        sys.exit("議題が1件もありません。")
    rows = []
    last = len(items) - 1
    for i, item in enumerate(items):
        missing = [k for k in AGENDA_REQUIRED_COLUMNS if not item.get(k)]
        if missing:
            sys.exit(f"議題の{i + 1}行目に不足があります: {', '.join(missing)}")

        sub_lines = ""
        note = (item.get("説明・論点") or "").strip()
        if note:
            sub_lines += SUB_LINE.format(label="", text=html.escape(note))
        material = (item.get("資料/備考") or "").strip()
        if material:
            sub_lines += SUB_LINE.format(label="資料/備考：", text=html.escape(material))

        minutes = str(item.get("想定時間（分）", "")).strip()
        if minutes.endswith("分"):
            minutes = minutes[:-1]

        rows.append(
            ROW_TEMPLATE.format(
                no=html.escape(item.get("No", "")),
                topic=html.escape(item.get("議題", "")),
                sub_lines=sub_lines,
                owner=html.escape(item.get("担当", "")),
                minutes=html.escape(minutes),
                border="" if i == last else ROW_BORDER,
            )
        )
    return "\n".join(rows)


def main():
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("--workbook", type=Path, help="会議情報・議題シートを含むExcelファイル(.xlsx)")
    parser.add_argument("--info", type=Path, help="会議情報CSVファイル（--items とセットで指定）")
    parser.add_argument("--items", type=Path, help="議題CSVファイル（--info とセットで指定）")
    parser.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE, help="元になるHTMLテンプレート")
    parser.add_argument("--output", type=Path, help="出力先HTMLファイル（省略時は output/agenda_YYYY-MM.html）")
    args = parser.parse_args()

    print(f"generate_agenda_email.py version {SCRIPT_VERSION}")

    if args.workbook:
        info, items = read_workbook(args.workbook)
    elif args.info and args.items:
        info, items = read_info_csv(args.info), read_items_csv(args.items)
    else:
        parser.error("--workbook か、--info と --items の組み合わせのいずれかを指定してください。")

    info = normalize_info_keys(info)
    validate_info(info)
    month, day = parse_month_day(info["月日"])

    title = info.get("件名") or f"【{info['年']}年{month}月度】定例会議"

    if not args.template.exists():
        sys.exit(f"テンプレートが見つかりません: {args.template}")
    text = args.template.read_text(encoding="utf-8")
    text = replace_field(text, "TITLE", title)
    text = replace_field(text, "DATETIME", format_datetime(info, month, day))
    text = replace_field(text, "PLACE", info["場所"])
    text = replace_field(text, "CHAIR", info["議長"])
    text = replace_field(text, "RECORDER", info["記録係"])
    text = replace_field(text, "ATTENDEES", info["出席者"])
    text = replace_field(text, "ABSENTEES", info.get("欠席者") or "なし")
    text = replace_rows(text, "ROWS:AGENDA", build_agenda_rows(items))

    output = args.output
    if output is None:
        output = REPO_ROOT / "output" / f"agenda_{info['年']}-{str(month).zfill(2)}.html"
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(text, encoding="utf-8")
    print(f"生成しました: {output}")


if __name__ == "__main__":
    main()
