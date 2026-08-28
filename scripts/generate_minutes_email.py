#!/usr/bin/env python3
"""スプレッドシート（Excel/CSV）の会議情報・審議内容を議事録メールテンプレートに流し込む。

使い方:
    # Excelワークブック（シート「会議情報」「議事録」「決定事項」「アクションアイテム」を含む .xlsx）から生成
    python3 scripts/generate_minutes_email.py --workbook meeting_data.xlsx

    # CSVファイル（会議情報・議事録・決定事項・アクションアイテムを分けて管理）から生成
    python3 scripts/generate_minutes_email.py \
        --info meeting_info.csv \
        --minutes minutes_items.csv \
        --decisions decisions.csv \
        --actions action_items.csv

出力先は省略時 output/<YYYYMMDD>_meeting_minutes.html。
ブラウザで開いて全選択→コピーし、メールソフトの本文に貼り付けて送信する。

入力フォーマットは scripts/sample_data/ のサンプルを参照。
「議事録」シートの行数（議事の数）は毎回変わってよい。1件でも100件でも、
シートの行を増減させればそのままメール本文の件数に反映される。
「決定事項」「アクションアイテム」も同様に件数自由（0件なら「特になし」と表示）。
"""
import argparse
import csv
import datetime
import html
import re
import sys
import unicodedata
from pathlib import Path

SCRIPT_VERSION = "2026-08-28a"

REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_TEMPLATE = REPO_ROOT / "meeting_minutes_template.html"

REQUIRED_INFO_KEYS = ["年", "月日", "時間帯", "場所", "議長", "記録係", "出席者"]
OPTIONAL_INFO_KEYS = [
    "欠席者",
    "件名",
    "回答期限",
    "次回年",
    "次回月日",
    "次回時間帯",
    "次回場所",
    "次回議題",
]
# 「月日」の項目名として区切り文字だけ違う表記（月/日、月ノ日、月・日 など）を許容する。
MONTH_DAY_SEPARATORS = "/\\-ノ・　 "

# 「議事録」シートの列名。議事の数は自由（行を増減すればそのまま反映される）。
MINUTES_REQUIRED_COLUMNS = ["No", "議題"]
MINUTES_OPTIONAL_COLUMNS = ["報告内容", "議論・意見", "結論"]
MINUTES_COLUMNS = MINUTES_REQUIRED_COLUMNS + MINUTES_OPTIONAL_COLUMNS

# 「決定事項」シートの列名。
DECISION_REQUIRED_COLUMNS = ["内容"]
DECISION_OPTIONAL_COLUMNS = ["No"]
DECISION_COLUMNS = DECISION_OPTIONAL_COLUMNS + DECISION_REQUIRED_COLUMNS

# 「アクションアイテム」シートの列名。「期限」「状況」は空欄可（状況未入力は「未着手」扱い）。
ACTION_REQUIRED_COLUMNS = ["タスク内容", "担当"]
ACTION_OPTIONAL_COLUMNS = ["No", "期限", "状況"]
ACTION_COLUMNS = ACTION_OPTIONAL_COLUMNS + ACTION_REQUIRED_COLUMNS

STATUS_BADGES = {
    "未着手": ("#b5750c", "#fdf1dc"),
    "対応中": ("#0c447c", "#e6f1fb"),
    "完了": ("#27500a", "#eaf3de"),
}
DEFAULT_STATUS = "未着手"

WEEKDAY_JA = ["月", "火", "水", "木", "金", "土", "日"]

MINUTES_ITEM_TEMPLATE = """  <tr>
    <td style="padding:0 40px 8px 40px;" class="fluid-pad">
      <table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" style="border:1px solid #d9dee4; border-radius:6px;">
        <tr>
          <td style="background-color:#f4f6f8; padding:10px 18px; font-size:13px; font-weight:bold; color:#1a1a1a; border-bottom:1px solid #d9dee4;">
            議題{no}：{topic}
          </td>
        </tr>
        <tr>
          <td style="padding:14px 18px 16px 18px; font-size:12.5px; color:#333333; line-height:1.8;">
            {body}
          </td>
        </tr>
      </table>
    </td>
  </tr>"""
MINUTES_BODY_LINE = "<strong>{label}：</strong>{text}"

DECISION_ROW_TEMPLATE = """              <tr>
                <td style="font-size:12.5px; color:#2a2a2a; line-height:1.8; padding:2px 0;">{no}. {content}</td>
              </tr>"""
DECISION_EMPTY = """              <tr>
                <td style="font-size:12.5px; color:#2a2a2a; line-height:1.8; padding:2px 0;">特になし</td>
              </tr>"""

ACTION_ROW_TEMPLATE = """        <tr>
          <td align="center" style="font-size:12.5px; color:#1a1a1a; padding:11px 6px;{border}">{no}</td>
          <td style="font-size:12.5px; color:#1a1a1a; padding:11px 10px;{border}">{task}</td>
          <td style="font-size:12.5px; color:#1a1a1a; padding:11px 10px;{border}">{owner}</td>
          <td align="center" style="font-size:12.5px; color:#1a1a1a; padding:11px 6px;{border}">{due}</td>
          <td align="center" style="padding:11px 6px;{border}">
            <span style="display:inline-block; font-size:11px; color:{fg}; background-color:{bg}; border-radius:10px; padding:2px 10px;">{status}</span>
          </td>
        </tr>"""
ACTION_ROW_BORDER = " border-bottom:1px solid #eef0f2;"
ACTION_EMPTY = """        <tr>
          <td colspan="5" align="center" style="font-size:12.5px; color:#8a92a0; padding:14px 6px;">特になし</td>
        </tr>"""


def normalize_column(text):
    """列見出しの全角/半角括弧・空白の違いを吸収して比較できる形にする。"""
    text = unicodedata.normalize("NFKC", str(text)).strip()
    for ch in "（）() 　":
        text = text.replace(ch, "")
    return text


def make_record_mapper(columns):
    lookup = {normalize_column(c): c for c in columns}

    def mapper(raw):
        mapped = {}
        for key, value in raw.items():
            canonical = lookup.get(normalize_column(key))
            if canonical:
                mapped[canonical] = value
        return {
            k: (str(mapped[k]).strip() if mapped.get(k) is not None else "")
            for k in columns
        }

    return mapper


map_minutes_record = make_record_mapper(MINUTES_COLUMNS)
map_decision_record = make_record_mapper(DECISION_COLUMNS)
map_action_record = make_record_mapper(ACTION_COLUMNS)


def cell_to_str(value):
    """Excelセルの値を文字列化する。

    - 整数値扱いの小数（5.0など）は末尾の.0を除く。
    - 「8/28」のように入力してExcelが自動で日付型に変換したセルは
      "8/28" 形式のテキストに戻す（parse_month_dayが読める形にする）。
    """
    if value is None:
        return ""
    if isinstance(value, (datetime.datetime, datetime.date)):
        return f"{value.month}/{value.day}"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def read_info_csv(path):
    info = {}
    with open(path, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if not row.get("項目"):
                continue
            info[row["項目"].strip()] = (row.get("内容") or "").strip()
    return info


def read_records_csv(path, mapper, columns):
    records = []
    with open(path, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            record = mapper(row)
            if not any(record.get(k) for k in columns):
                continue
            records.append(record)
    return records


def find_header_row(ws, first_column_name, scan_limit=20):
    """A列が first_column_name と一致する行を見出し行として探す。"""
    limit = min(scan_limit, ws.max_row)
    for row_num, row in enumerate(
        ws.iter_rows(min_row=1, max_row=limit, values_only=True), start=1
    ):
        if row and unicodedata.normalize("NFKC", cell_to_str(row[0])).strip().lower() == first_column_name.lower():
            return row_num, [cell_to_str(c) for c in row]
    return None, None


def read_sheet_records(ws, first_column_names, mapper, columns):
    """first_column_names に列挙した候補のいずれかがA列に現れる行を見出し行とみなす。"""
    if isinstance(first_column_names, str):
        first_column_names = [first_column_names]
    header_row_index = header = None
    for name in first_column_names:
        header_row_index, header = find_header_row(ws, name)
        if header is not None:
            break
    if header is None:
        candidates = "」「".join(first_column_names)
        sys.exit(
            f'シート「{ws.title}」で見出し行（A列が「{candidates}」のいずれかの行）が見つかりません。'
        )
    records = []
    for row in ws.iter_rows(min_row=header_row_index + 1, values_only=True):
        if not row or all(v is None for v in row):
            continue
        raw = {header[i]: cell_to_str(row[i]) for i in range(min(len(header), len(row)))}
        record = mapper(raw)
        if not any(record.get(k) for k in columns):
            continue
        records.append(record)
    return records


def read_workbook(path):
    try:
        import openpyxl
    except ImportError:
        sys.exit(
            "xlsxファイルを読むには openpyxl が必要です。`pip install openpyxl` を実行してください。"
        )
    wb = openpyxl.load_workbook(path, data_only=True)
    if "会議情報" not in wb.sheetnames:
        sys.exit('シート「会議情報」が見つかりません。')
    if "議事録" not in wb.sheetnames:
        sys.exit('シート「議事録」が見つかりません。')

    info_ws = wb["会議情報"]
    info = {}
    for row in info_ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        key, value = row[0], row[1] if len(row) > 1 else None
        info[cell_to_str(key)] = cell_to_str(value)

    minutes = read_sheet_records(wb["議事録"], "No", map_minutes_record, MINUTES_COLUMNS)

    decisions = []
    if "決定事項" in wb.sheetnames:
        decisions = read_sheet_records(
            wb["決定事項"], ["No", "内容"], map_decision_record, DECISION_COLUMNS
        )

    actions = []
    if "アクションアイテム" in wb.sheetnames:
        actions = read_sheet_records(
            wb["アクションアイテム"], "No", map_action_record, ACTION_COLUMNS
        )

    return info, minutes, decisions, actions


def strip_separators(text):
    for ch in MONTH_DAY_SEPARATORS:
        text = text.replace(ch, "")
    return text


def normalize_info_keys(info):
    """「月/日」「月ノ日」「月」のように書き方が違う項目名を「月日」に寄せる。"""
    if "月日" not in info:
        for key in list(info):
            if strip_separators(key) == "月日":
                info["月日"] = info[key]
                break
        else:
            if "月" in info:
                info["月日"] = info["月"]
    return info


def validate_info(info):
    missing = [k for k in REQUIRED_INFO_KEYS if not info.get(k)]
    if missing:
        found = ", ".join(repr(k) for k in info.keys()) or "(1件も読み取れませんでした)"
        sys.exit(
            f"会議情報に不足があります: {', '.join(missing)}\n"
            f"読み取れた項目名: {found}"
        )


MONTH_DAY_VALUE_RE = re.compile(r"(\d{1,2})\s*[/\-月]\s*(\d{1,2})")


def parse_month_day(value):
    """「9/10」「9-10」「9月10日」など、月日の数字が読み取れる形式から (月, 日) を取り出す。"""
    text = str(value).strip()
    m = MONTH_DAY_VALUE_RE.search(text)
    if m:
        return int(m.group(1)), int(m.group(2))
    sys.exit(
        f"「月日」の値が正しくありません: {value!r}。"
        "「9/10」または「9月10日」のように月と日の数字を入力してください。"
    )


def format_datetime(year, month, day, time_range):
    try:
        d = datetime.date(int(year), month, day)
    except (TypeError, ValueError):
        sys.exit(f"「年」の値が正しくありません（年={year!r}）。")
    weekday = WEEKDAY_JA[d.weekday()]
    return f"{year}年{month}月{day}日（{weekday}）　{time_range}"


def replace_field(text, key, value):
    pattern = re.compile(r"(<!--F:%s-->)(.*?)(<!--/F-->)" % re.escape(key), re.DOTALL)
    if not pattern.search(text):
        sys.exit(f"テンプレートにマーカー <!--F:{key}--> が見つかりません。")
    return pattern.sub(lambda m: m.group(1) + html.escape(value) + m.group(3), text)


def replace_rows(text, marker, rows_html):
    pattern = re.compile(
        r"(<!--%s:START-->)(.*?)(<!--%s:END-->)" % (re.escape(marker), re.escape(marker)),
        re.DOTALL,
    )
    if not pattern.search(text):
        sys.exit(f"テンプレートにマーカー <!--{marker}:START--> が見つかりません。")
    return pattern.sub(lambda m: m.group(1) + "\n" + rows_html + "\n" + m.group(3), text)


def build_minutes_blocks(items):
    if not items:
        sys.exit("議事録（審議内容）が1件もありません。")
    blocks = []
    for i, item in enumerate(items):
        missing = [k for k in MINUTES_REQUIRED_COLUMNS if not item.get(k)]
        if missing:
            sys.exit(f"議事録の{i + 1}行目に不足があります: {', '.join(missing)}")

        lines = []
        for label, key in (("報告内容", "報告内容"), ("議論・意見", "議論・意見"), ("結論", "結論")):
            text = (item.get(key) or "").strip()
            if text:
                lines.append(MINUTES_BODY_LINE.format(label=label, text=html.escape(text)))
        body = "<br>\n            ".join(lines) if lines else "特記事項なし。"

        blocks.append(
            MINUTES_ITEM_TEMPLATE.format(
                no=html.escape(item.get("No", "") or str(i + 1)),
                topic=html.escape(item.get("議題", "")),
                body=body,
            )
        )
    return "\n".join(blocks)


def build_decision_rows(decisions):
    if not decisions:
        return DECISION_EMPTY
    rows = []
    for i, item in enumerate(decisions):
        content = (item.get("内容") or "").strip()
        if not content:
            continue
        no = item.get("No") or str(i + 1)
        rows.append(DECISION_ROW_TEMPLATE.format(no=html.escape(no), content=html.escape(content)))
    return "\n".join(rows) if rows else DECISION_EMPTY


def build_action_rows(actions):
    if not actions:
        return ACTION_EMPTY
    rows = []
    last = len(actions) - 1
    for i, item in enumerate(actions):
        missing = [k for k in ACTION_REQUIRED_COLUMNS if not item.get(k)]
        if missing:
            sys.exit(f"アクションアイテムの{i + 1}行目に不足があります: {', '.join(missing)}")
        status = (item.get("状況") or "").strip() or DEFAULT_STATUS
        fg, bg = STATUS_BADGES.get(status, ("#5a6472", "#eef0f2"))
        rows.append(
            ACTION_ROW_TEMPLATE.format(
                no=html.escape(item.get("No") or str(i + 1)),
                task=html.escape(item.get("タスク内容", "")),
                owner=html.escape(item.get("担当", "")),
                due=html.escape(item.get("期限", "") or "-"),
                status=html.escape(status),
                fg=fg,
                bg=bg,
                border="" if i == last else ACTION_ROW_BORDER,
            )
        )
    return "\n".join(rows)


def main():
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("--workbook", type=Path, help="会議情報・議事録などのシートを含むExcelファイル(.xlsx)")
    parser.add_argument("--info", type=Path, help="会議情報CSVファイル")
    parser.add_argument("--minutes", type=Path, help="議事録（審議内容）CSVファイル（--info とセットで指定）")
    parser.add_argument("--decisions", type=Path, help="決定事項CSVファイル（省略可）")
    parser.add_argument("--actions", type=Path, help="アクションアイテムCSVファイル（省略可）")
    parser.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE, help="元になるHTMLテンプレート")
    parser.add_argument(
        "--output", type=Path, help="出力先HTMLファイル（省略時は output/<YYYYMMDD>_meeting_minutes.html）"
    )
    args = parser.parse_args()

    print(f"generate_minutes_email.py version {SCRIPT_VERSION}")

    if args.workbook:
        info, minutes, decisions, actions = read_workbook(args.workbook)
    elif args.info and args.minutes:
        info = read_info_csv(args.info)
        minutes = read_records_csv(args.minutes, map_minutes_record, MINUTES_COLUMNS)
        decisions = (
            read_records_csv(args.decisions, map_decision_record, DECISION_COLUMNS)
            if args.decisions
            else []
        )
        actions = (
            read_records_csv(args.actions, map_action_record, ACTION_COLUMNS)
            if args.actions
            else []
        )
    else:
        parser.error("--workbook か、--info と --minutes の組み合わせのいずれかを指定してください。")

    info = normalize_info_keys(info)
    validate_info(info)
    month, day = parse_month_day(info["月日"])

    title = info.get("件名") or f"【{info['年']}年{month}月度】定例会議 議事録"
    reply_deadline = info.get("回答期限") or "追ってご案内する期日"

    next_datetime = ""
    if info.get("次回年") and info.get("次回月日"):
        next_month, next_day = parse_month_day(info["次回月日"])
        next_datetime = format_datetime(
            info["次回年"], next_month, next_day, info.get("次回時間帯") or "時間未定"
        )
    next_place = info.get("次回場所") or "未定"
    next_agenda = info.get("次回議題") or "未定"

    if not args.template.exists():
        sys.exit(f"テンプレートが見つかりません: {args.template}")
    text = args.template.read_text(encoding="utf-8")
    text = replace_field(text, "TITLE", title)
    text = replace_field(text, "REPLY_DEADLINE", reply_deadline)
    text = replace_field(text, "DATETIME", format_datetime(info["年"], month, day, info["時間帯"]))
    text = replace_field(text, "PLACE", info["場所"])
    text = replace_field(text, "CHAIR", info["議長"])
    text = replace_field(text, "RECORDER", info["記録係"])
    text = replace_field(text, "ATTENDEES", info["出席者"])
    text = replace_field(text, "ABSENTEES", info.get("欠席者") or "なし")
    text = replace_field(text, "NEXT_DATETIME", next_datetime or "未定")
    text = replace_field(text, "NEXT_PLACE", next_place)
    text = replace_field(text, "NEXT_AGENDA", next_agenda)
    text = replace_rows(text, "ROWS:DECISIONS", build_decision_rows(decisions))
    text = replace_rows(text, "ROWS:MINUTES", build_minutes_blocks(minutes))
    text = replace_rows(text, "ROWS:ACTIONS", build_action_rows(actions))

    output = args.output
    if output is None:
        date_prefix = f"{info['年']}{str(month).zfill(2)}{str(day).zfill(2)}"
        output = REPO_ROOT / "output" / f"{date_prefix}_meeting_minutes.html"
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(text, encoding="utf-8")
    print(f"生成しました: {output}")


if __name__ == "__main__":
    main()
